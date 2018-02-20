from openpyxl import Workbook
from openpyxl.styles import Font

class Excel():

    """Class to work with Excel files"""

    def __init__(self, filename, worksheetTitle):
        self.filename = filename
        self.workbook = Workbook()

        # one worksheet is created by default so add it to dictionary of
        # worksheets and assign its title
        self.worksheets = {worksheetTitle: self.workbook.active}
        self.worksheets[worksheetTitle].title = worksheetTitle
        self.styles = {'bold': Font(bold=True), 'non-bold': Font(bold=False)}


    def writeFile(self):
        self.workbook.save(filename=self.filename)

    def addWorksheet(self, worksheetTitle):
        self.worksheets[worksheetTitle] = self.workbook.create_sheet(worksheetTitle)


    def writeRow(self, worksheet, startRow, startColumn, bold, data):
        """Write a tuple of 'data' to a given worksheet at a certain row"""
        """and column"""

        ws = self.workbook[worksheet]

        for columnIndex, column in enumerate(data):
            ws.cell(row=startRow, column=startColumn + columnIndex).value = column

            if bold:
                ws.cell(row=startRow , column=startColumn + columnIndex).font = self.styles['bold']
            else:
                ws.cell(row=startRow, column=startColumn + columnIndex).font = self.styles['non-bold']


    def writeRows(self, worksheet, startRow, startColumn, bold, data):
        """Write a 2-dimensional list of 'data' to a given worksheet,"""
        """starting at a given row and column"""

        ws = self.workbook[worksheet]


        for rowIndex, row in enumerate(data):
            for columnIndex, column in enumerate(row):
                ws.cell(row=startRow + rowIndex, column=startColumn + columnIndex).value = column

                if bold:
                    ws.cell(row=startRow + rowIndex, column=startColumn + columnIndex).font = self.styles['bold']
                else:
                    ws.cell(row=startRow + rowIndex, column=startColumn + columnIndex).font = self.styles['non-bold']
